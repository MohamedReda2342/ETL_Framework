
import openpyxl
import pandas as pd
import xlrd
from collections import OrderedDict
import json
from colorama import init, Fore, Back, Style
init(autoreset=True)  # Initialize colorama

import excel2json

def wb_to_jjson(wbfile):
    wb = xlrd.open_workbook(wbfile)
    sh = wb.sheet_by_index(0)
    data_list = []
    for rownum in range(1, sh.nrows):
     data = OrderedDict()

    row_values = sh.row_values(rownum)
    data['<Column Name1>'] = row_values[0]
    data['<Column Name2>'] = row_values[1]
    data_list.append(data)

    with open("RulesJson.json", "w", encoding="utf-8") as writeJsonfile:
      json.dump(data_list, writeJsonfile, indent=4,default=str) 
    return True


def excel_json(filename):
   print(filename)
   smx_json= excel2json.convert_from_file(filename)
   print(Fore.YELLOW+'excel2json')
   print(smx_json)
   print(type(smx_json))
   return True




def load_smx_file(filename):
    work_book = openpyxl.load_workbook(filename)
    worksheets= work_book.worksheets
    sheetnames=work_book.sheetnames
    #print(sheetnames)
    
    smx_model={}

    for work_sheet in sheetnames:
        df = pd.read_excel(filename,work_sheet)
        cols = [x.lower() for x in df.columns]
        df.columns=cols
        smx_model[work_sheet.lower()] = df #list(df.columns)

    smx_df = pd.DataFrame(smx_model.items())
    return smx_model, smx_df


def save_json(smx_data):

# Define the JSON object

    # Specify the file path
    file_path = 'smx_data_dict_7.json'
    print(Fore.YELLOW+"in save and the inut file type is :")
    print(type(smx_data))
    for k in smx_data.keys():
       df=smx_data[k]
       smx_data[k]=df.to_json(orient='records') 
    print(smx_data)
    #smx_data_json=json.dump(smx_data,  indent=4)
    #print(smx_data_json)
    # Write the JSON object to the file
    
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(smx_data, f)
    print(f"Data saved to {file_path}")



if __name__ == '__main__':

   
    print('=====================================================================')

    #data_file = 'C:/Users/hanaa.hamad/Documents/projects/eBANK/Preparation/ETL_script_generator/WAVZ_SMX_Version_0.5.xlsx'
    #data_file = 'WAVZ_SMX_Version_0.5.xlsx'
    data_file = 'WAVZ_SMX_Version_0.6 060725.xlsx'
    data_file = 'C:/Workspaces/HH-ETL-scripts-main/documentation/scripts/V7/WAVZ_SMX_Version_0.7.xlsx'
    smx_model, smx_df=load_smx_file(data_file)
    smx_json_obj= smx_df.to_json(orient='records') 
    #smx_data=json.load(smx_json_obj)
    print(type(smx_model))
    print(type(smx_df))

    #print(type(smx_json))
    #save_json(smx_json)
    save_json(smx_model)

    #excel_json(data_file)



