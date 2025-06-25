import streamlit as st
import openpyxl
import pandas as pd
# import util.script_generator as sg
import util.df_utlis as dfutils
from util.auth import check_authentication
import difflib
# Check authentication - this must be the first Streamlit command
authenticator = check_authentication()
# Only authenticated users will see content below this point
st.markdown(
    """<style>
        .element-container:nth-of-type(3) button {
            height: 3em;
        }
        </style>""",
    unsafe_allow_html=True,
)

# Set page config
# st.set_page_config(page_title="WAVZ SMX Workbook Comparator", layout="wide")


# Function to load and process workbook
def load_workbook(uploaded_file):
    if uploaded_file is None:
        return None
    
    # Read all sheets from the Excel file
    xls = pd.ExcelFile(uploaded_file)
    sheets = {}
    
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_name, header=0)
        # Clean up the dataframe (remove empty rows/columns)
        df = df.dropna(how='all').dropna(axis=1, how='all')
        sheets[sheet_name] = df
    
    return sheets

# Function to compare two dataframes
def compare_dataframes(df1, df2, sheet_name):
    if df1 is None or df2 is None:
        return None
    
    # Convert dataframes to strings line by line
    df1_lines = df1.to_string(index=False).split('\n')
    df2_lines = df2.to_string(index=False).split('\n')
    
    # Generate diff
    diff = difflib.unified_diff(
        df1_lines, 
        df2_lines,
        fromfile=f"Workbook 1 - {sheet_name}",
        tofile=f"Workbook 2 - {sheet_name}",
        lineterm=""
    )
    
    return '\n'.join(diff)

# Function to highlight differences between two dataframes
def highlight_diff(df1, df2):
    if df1 is None or df2 is None:
        return None
    
    # Ensure dataframes have same shape and columns
    if not df1.shape == df2.shape or not all(df1.columns == df2.columns):
        return pd.DataFrame(["Dataframes have different structures. Cannot compare cell by cell."])
    
    # Create a styled dataframe that highlights differences
    comparison_df = df1.copy()
    for col in df1.columns:
        for idx in df1.index:
            val1 = str(df1.at[idx, col])
            val2 = str(df2.at[idx, col])
            if val1 != val2:
                comparison_df.at[idx, col] = f"{val1} → {val2}"
    
    return comparison_df

# Function to find different rows
def compare_rows(df1, df2):
    if df1 is None or df2 is None:
        return None, None, None

    OuterJoinedDF = pd.merge(df1, df2, how="outer", indicator="Exist")
    only_in_df1 = OuterJoinedDF.query("Exist == 'left_only'")
    only_in_df2 = OuterJoinedDF.query("Exist == 'right_only'")
    different_rows_df = OuterJoinedDF.query("Exist != 'both'")
    return only_in_df1, only_in_df2, different_rows_df

# Function to compare full workbooks
def compare_full_workbooks(wb1_sheets, wb2_sheets):
    """Compare all sheets across both workbooks and return summary statistics"""
    comparison_summary = []
    
    # Get all unique sheet names
    all_sheets = set(wb1_sheets.keys()) | set(wb2_sheets.keys())
    common_sheets = set(wb1_sheets.keys()) & set(wb2_sheets.keys())
    
    for sheet_name in sorted(all_sheets):
        sheet_summary = {
            'Sheet': sheet_name,
            'Status': '',
            'Rows Only in WB1': 0,
            'Rows Only in WB2': 0,
            'Rows with Differences': 0
        }
        
        if sheet_name not in wb1_sheets:
            sheet_summary['Status'] = 'Only in Workbook 2'
        elif sheet_name not in wb2_sheets:
            sheet_summary['Status'] = 'Only in Workbook 1'
        else:
            # Sheet exists in both workbooks
            df1 = wb1_sheets[sheet_name]
            df2 = wb2_sheets[sheet_name]

            if  len(df1.columns) > 0 and len(df2.columns) > 0:
                only_in_df1, only_in_df2, different_rows = compare_rows(df1, df2)
                
                sheet_summary['Rows Only in WB1'] = len(only_in_df1) 
                sheet_summary['Rows Only in WB2'] = len(only_in_df2) 
                sheet_summary['Rows with Differences'] = len(different_rows) if different_rows is not None else 0
                
                total_sheet_diffs = sheet_summary['Rows Only in WB1'] + sheet_summary['Rows Only in WB2'] + sheet_summary['Rows with Differences']
                
                if total_sheet_diffs == 0:
                    sheet_summary['Status'] = 'Identical'
                else:
                    sheet_summary['Status'] = 'Has Differences'
            else:
                sheet_summary['Status'] = 'Cannot Compare'
        
        comparison_summary.append(sheet_summary)
    
    return pd.DataFrame(comparison_summary)

# Upload files
col1, col2 = st.columns(2)
with col1:
    st.subheader("Workbook 1")
    workbook1 = st.file_uploader("Upload first workbook", type=["xlsx"], key="wb1")
    
with col2:
    st.subheader("Workbook 2")
    workbook2 = st.file_uploader("Upload second workbook", type=["xlsx"], key="wb2")

# Load workbooks if files are uploaded
wb1_sheets = load_workbook(workbook1)
wb2_sheets = load_workbook(workbook2)

if wb1_sheets and wb2_sheets:
    # Generate full workbook comparison
    with st.spinner("Analyzing all sheets..."):
        summary_df = compare_full_workbooks(wb1_sheets, wb2_sheets)
    
    # Get sheets with differences for drill down
    sheets_with_differences = summary_df[summary_df['Status'] == 'Has Differences']
    
    # Show unique sheets
    unique_sheets = summary_df[summary_df['Status'].isin(['Only in Workbook 1', 'Only in Workbook 2'])]
    if not unique_sheets.empty:
        st.subheader("📄 Sheets Present in Only One Workbook")
        st.dataframe(unique_sheets, use_container_width=True)
    
    # Option to drill down into specific sheets
    if not sheets_with_differences.empty:
        st.subheader("🔍 Drill Down Analysis")
        selected_sheet_for_detail = st.selectbox(
            "Select a sheet with differences for detailed analysis:",
            sheets_with_differences['Sheet'].tolist()
        )
        
        if selected_sheet_for_detail:
            st.write(f"Detailed analysis for: **{selected_sheet_for_detail}**")
            df1 = wb1_sheets[selected_sheet_for_detail]
            df2 = wb2_sheets[selected_sheet_for_detail]
                        
            only_in_df1, only_in_df2, different_rows = compare_rows(df1, df2)
            
            # Show the differences
            if not only_in_df1.empty:
                st.write(f"Rows only in Workbook 1")
                st.dataframe(only_in_df1)
            
            if not only_in_df2.empty:
                st.write(f"Rows only in Workbook 2")
                st.dataframe(only_in_df2)
            
            if not different_rows.empty:
                st.write(f"**Rows with differences:** {len(different_rows)}")
                st.dataframe(different_rows)
            
elif wb1_sheets or wb2_sheets:
    st.warning("Please upload both workbooks to compare")
else:
    st.info("Upload two WAVZ SMX workbooks to compare them")
   
        # Show logout button
with st.sidebar:
    authenticator.logout()