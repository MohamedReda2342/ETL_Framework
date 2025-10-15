import streamlit as st
import pandas as pd
import openpyxl
from dotenv import load_dotenv
import os
import read_env
import numpy as np
from read_env import CurrentEnv 
import ast 
import sys
from colorama import init, Fore, Back, Style
init(autoreset=True)  # Initialize colorama
from itertools import chain  
# import regex as re
import psutil
import os
from util import Queries 
from util import df_utlis

def load_script_model():
    obj = pd.read_pickle(r'pickled_df/bkey_functions.pkl')
    return obj

def create_template(filename, key_type):
    functions_DF = pd.read_excel(filename,'FUNCTIONS')
    #functions_DF['function_name'] = 
    fs= list(functions_DF['Syntax'])
    start='.'
    end='('
    fs= [s.strip('EXEC') for s in fs]
    functions = [s[s.find(start)+len(start):s.rfind(end)] for s in fs]

    return functions 


def load_smx_file(filename):
    work_book = openpyxl.load_workbook(data_file)
    worksheets= work_book.worksheets
    sheetnames=work_book.sheetnames
    #print(sheetnames)
    
    smx_model={}

    for work_sheet in sheetnames:
        df = pd.read_excel(filename,work_sheet)
        cols = [x.lower() for x in df.columns]
        df.columns=cols
        smx_model[work_sheet.lower()] = df

    smx_df = pd.DataFrame(smx_model.items())
    return smx_model, smx_df


def load_syntax_model(syntax_file):
    work_book = openpyxl.load_workbook(syntax_file)
    worksheets= work_book.worksheets
    sheetnames=work_book.sheetnames
    print('syntax model sheet names')
    print(sheetnames)
    
    script_model={}

    for work_sheet in sheetnames:
        df = pd.read_excel(syntax_file,work_sheet)
        print(df)
        #print(df)
        cols = [x.lower() for x in df.columns]
        df.columns=cols
        script_model[work_sheet.lower()] = df #list(df.columns)

    script_df = pd.DataFrame(script_model.items())
    print('=====================================================================')
    #print('script model')
    print('=====================================================================')
    print('=====================================================================')
    #print('script df')
    print('=====================================================================')
    #print(script_df)
    return script_model #, script_df


def load_env():
    env=read_env.CurrentEnv()
    print(env)
    return env

def filter_key_type(script_dict, key_type):
    print('================     filter_key_type      =====================================================')
    print('========= returns a filtered script model to the key_type selected  ============')

    #print(script_dict)
    for k in script_dict.keys():
        print (k)
        
    df=script_dict['functions']
    #print('=============================================== df',df)
    filtered_functions = df[df['key_type'] ==key_type]
    
    print('========= now iterate over the filtered functions so we can get the parameters  ============')
    
    df2=script_dict['parameters']
    df_filtered = pd.merge(filtered_functions, df2, on='function_code')
    #print(df_filtered)
    df3=script_dict['unique_parameters']
    df_filtered2=  pd.merge(df_filtered, df3, left_on='parameters', right_on='parameter_name')
    #print(df_filtered2)


    return df_filtered2





def get_env_dict(env, bi_flag):
    envvar = CurrentEnv(env, bi_flag )
    env_attributes = {
        k: v
        for k, v in vars(envvar).items()
        if not callable(v) and not k.startswith("__")
    }
    print(env_attributes)
    return env_attributes



def flatten(a):  
    res = []  
    for x in a:  
        if isinstance(x, list):  
            res.extend(flatten(x))  
        else:  
            res.append(x)  
    return res  

def join_bkey_stg_stream(smx_tabs, df,smx_dict):
    print(Back.LIGHTGREEN_EX+ "==================    join_bkey_stg_stream   ========================================================")
    print(df)
    print(df.columns)
    print(smx_tabs)
    #BKEY and STG tables
    df1=smx_dict['bkey']
    df2=smx_dict['stg tables']
    on_cols=['key set name','key domain name']
    merged= pd.merge(df1, df2, on=on_cols, how='left')
    print(Fore.LIGHTMAGENTA_EX+"merged bkey and stg tables")
    print(merged)
    df3=smx_dict['stream']
    print(df3)
    #hard coded process type = 21 ==> _bkey
    search_string='_BKEY'
    #search_string=''
    filtered_df = df3[df3['stream name'].str.contains(search_string)]
    print(Fore.LIGHTMAGENTA_EX+"filtered stream on _bkey")
    print(filtered_df)
    #on_cols=[source system alias, system name]
    smx_df = pd.merge(merged, filtered_df, left_on='source system alias', right_on='system name', how='left')
    print(smx_df)
    print(Fore.LIGHTMAGENTA_EX+"filtered stream on _bkey")

    print(smx_df.columns)
    return smx_df

def join_columns(smx_tabs, df,smx_dict, key_type):
    """
    This functions will perform the following:
    1) join tabs (dataframes) from the smx workbook based on the key_type
    2) perform some modifications on the dataframe:
        - reading as str type
        - for the core process, we need to replace the operation by process type

    """
    
    print(Fore.YELLOW+"-------------------- In the join function -----------------------")
    the_joined_smx_tabs= pd.DataFrame()
    #print(df)
    print(smx_tabs)
    print(key_type)
    if key_type=='EXEC_SRCI':
        search_string='_SRCI'
        df2=smx_dict['stg tables']
        
        df3=smx_dict['stream']
        filtered_df=df3[df3['stream name'].str.contains(search_string)]
        the_joined_smx_tabs = pd.merge(df2, filtered_df, left_on='source system alias', right_on='system name', how='left')
    elif key_type=='REG_CORE_PROCESS':
        print(Fore.LIGHTRED_EX+"*****************************************  **************************************")
        search_string='_CORE'
        df1=smx_dict['table mapping']
        print(df1.columns)
        print(df1['historization algorithm'])
        hist_algo =  df1['historization algorithm']
        hist_algo_dict = {'INSERT': [25,1] ,  'UPSERT': [23,0], 'HISTORY': [29,0]}
        hist_algo = list(map(lambda x: hist_algo_dict[x][0] , list(df1['historization algorithm'])))
        #hist_algo = list(map(lambda x: hist_algo_dict[x] , list(df1['historization algorithm'])))
        print(hist_algo)
        #df1['historization algorithm']=hist_algo
        df1['core process type'] = hist_algo
        #df Verification_Flag = 1
        hist_algo = list(map(lambda x: hist_algo_dict[x][1] , list(df1['historization algorithm'])))
        df1['verification flag'] = hist_algo

        df2=smx_dict['stg tables']
        df3=smx_dict['stream']
        print(df3)
        df3_strings= df3.astype(str)
        print(df3_strings)
        filtered_df=df3[df3['stream name'].str.contains(search_string)]
        print(filtered_df)
        filtered_df= filtered_df.astype(str)
        df1_df2_merged= pd.merge(df1, df2, left_on='main source', right_on='table name stg', how='left')
        print(df1_df2_merged)
        the_joined_smx_tabs = pd.merge(df1_df2_merged, filtered_df, left_on='source system alias', right_on='system name', how='left')

    else:
       
       the_joined_smx_tabs= join_bkey_stg_stream(smx_tabs, df,smx_dict)

    print(Fore.YELLOW+'-------------- the returned joined df ---- ')
    print(the_joined_smx_tabs)
    print(the_joined_smx_tabs.columns)
    return the_joined_smx_tabs







def get_params_values_better(smx_tab, df, smx_dict, key_type):
    print(Back.LIGHTRED_EX+ "==================    get_params_values_better   ========================================================")
    print(key_type)
    print(smx_tab)
    print(df.columns)
    print(Back.CYAN+ "incoming df parameters - content of parameters")
    print(df.columns)
    print(df[['parameter_name','source', 'parameter']])

    smx_lst=list(df['smx_column'])
    smx_tabs_lst=list(df['source'])
    #smx_join_key=list(df.query('join_key == join_key')['join_key'])
    
    print(Fore.LIGHTGREEN_EX+ f'{smx_lst}')
    print(Fore.RED+"this is the smx_tabs_lst")
    print(smx_tabs_lst)
    #print(smx_join_key)
    
    smx_source_list= [x.split(',') for x in smx_tabs_lst]

    print(smx_source_list)
    
    ll2 = [item for item in smx_source_list if item != 'env']
    print(Back.CYAN+ "ll2")
    print(ll2)
    flattened_list = [item for sublist in ll2 for item in sublist]
    print(Back.CYAN+ "flattened_list")
    print(flattened_list)
    print("&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&")
    if 'env' in flattened_list:
        flattened_list.remove('env')
    print(Back.CYAN+ "flattened_list")
    print(flattened_list)

    multiple_source_list = [x for x in smx_source_list if len(x) > 1] #[ len(l) for l in smx_source_list]
    print(Back.CYAN+ "Multiple sources list")
    print(multiple_source_list)
    #check the multiple sources and the maximum in order to join the dataframes

    print(type(multiple_source_list))
    print(len(multiple_source_list))
    import itertools
    smx_df=pd.DataFrame()
    if len(multiple_source_list)>0:
        print(Fore.LIGHTGREEN_EX+'here in multiple source list \n checking on the tabs')
        multiple_source_list = list(itertools.chain(*multiple_source_list))
        print(Back.CYAN+ "FLATTENED Multiple sources list")
        print(multiple_source_list)
        print(Back.CYAN+ "SET FLATTENED Multiple sources list")
        print(set(multiple_source_list))
        print(list(set(multiple_source_list)))
        smx_tabs=multiple_source_list 
        print(Fore.CYAN+f"{smx_tabs}")
        print(len(smx_tabs))   
        smx_df=join_columns(smx_tabs, df, smx_dict, key_type)
    else: 
        print(Fore.LIGHTGREEN_EX+'here in checking if multiple source list: else ie not multiple source')
        smx_tabs=flattened_list

    

    #smx_cols_list = [x.split(',') for x in smx_lst]
    #print(smx_cols_list)
    smx_list = [x.lower() for x in smx_lst if pd.notnull(x)]
    print(Back.CYAN+ "SMX list - content of smx_column")

    print(smx_list)
    print(smx_tab)
    smx_tabs_df=[]
    #smx_tabs=smx_tab.split(',')

    print("**********************************************************************************************")
    print(smx_df)

    if not(multiple_source_list):
        print(Fore.YELLOW+ " it is a single source SO WE JUST FILTER THE TAB ON THE COLUMNS")
        print(smx_dict.keys())
        print(smx_tab)
        smx_df=smx_dict[smx_tab]
        print(smx_df)

    print(Back.LIGHTRED_EX+ "----------------------------------------------------------------------------")
    

    smx_col_list= [x.split(',') for x in smx_list]
    print(Back.CYAN+ "smx_col_list")
    print(smx_col_list)
    flattened_list = [element for sublist in smx_col_list for element in sublist]
    print(flattened_list)
    print(Back.LIGHTRED_EX+ "----------------------------------------------------------------------------")
    
    #smx_list=flattened_list
    print(smx_df)
    print(smx_df.columns)
    smx_list=flattened_list
    print(Back.LIGHTCYAN_EX+ "----------------------------------------------------------------------------")

    print(smx_list)
    print(smx_df.columns)
    print(smx_df[smx_list])

    print(Back.LIGHTRED_EX+ "----------------------------------------------------------------------------")
    '''
    New section is here
    we create the unique set of smx_list and use the dataframe
    then we loop over the parameters and assign the values to the new columns

    '''
    print(Back.YELLOW+ "*********************************************************************************")

    print(Fore.YELLOW+" &&&&&&&&&&&&&&&&&   we start here ")

    smx_list_set = list(set(smx_list))
    print(smx_list_set)
    print(len(smx_list_set))
    
    print(Back.YELLOW+" smx_df ")
    smx_df=smx_df[smx_list_set]
    print(smx_df.shape)
    print(smx_df)

    print(Back.YELLOW+" iterate over the df to see what to do ")
    print(df[['parameter_name', 'source', 'parameter', 'smx_column', 'env_variable']])
    print(df.columns)
    final_df=pd.DataFrame()
    n=smx_df.shape[0]
    print(Fore.CYAN+f'printing n = {n}' )
    #process_type=df['Process_type']
    #check if there is a process type in the list of parameters
    print(Fore.YELLOW+' hhhhhhhhhhhhhhhhhhhhhhh checking the parameter name for process')
    process_type_row = df.loc[df['parameter_name'] == 'Process_Type']
    print(process_type_row)
    process_type_row=df.loc[df['parameter_name'].str.contains('Process_Type')]
    
    print(Fore.YELLOW+"*************    PROCESS TYPE    ********************************************")

    print(Fore.YELLOW+"*********************************************************************")
    print(Fore.YELLOW+f'{process_type_row}')
    print(Fore.YELLOW+"*********************************************************************")
    print(type(process_type_row))
    
    print(Fore.YELLOW+ f'checking the process type exists: {process_type_row}')

    process_type_str='null'
    if not process_type_row.empty:

        process_type=process_type_row['parameter']
        print(process_type_row)
        print(Fore.LIGHTRED_EX+ f'process_type : {process_type}')
        print(Fore.LIGHTYELLOW_EX+ f'{process_type}')
        print(type(process_type))
        values_list = process_type.tolist()
        process_type_str =values_list[0]
        print(Fore.CYAN +f'{process_type_str}')
    
    for index, row in df.iterrows():
        print(Fore.YELLOW+"------------------------ checking the rows attributes -------------------")
        #print(row)
        #print(row['prefix'])
        #print(row['parameter'])
        '''
        checking if the source is env or the smx document
        if from env, we need to remove the quotes check if the value is a number or not so we put it as is or with quotes
        '''
        if row['source']=='env':
            print(Fore.YELLOW+f'here in the env parameter')
            #print(row['parameter'])
            my_string=row['parameter']
            s=my_string.replace('"', '')
            #print(s.isdigit())
            if s.isdigit():
                #new_string = my_string.replace('"', '')
                #print(new_string)
                final_df[row['parameter_name']]=[s]*n
            elif (s=='None'):
                print("s=None")
                final_df[row['parameter_name']]=''*n
                #pd.Series()



            else:
                my_string=my_string.replace('"', "'")
                final_df[row['parameter_name']]=[my_string]*n
            
            col_values=[row['parameter']]*n
            print(Fore.LIGHTMAGENTA_EX+ "printing the environment values")
            #print(col_values)
            
           

        else:
            smx_cols= row['parameter'].split(',')
            print(Fore.CYAN+f'smx_cols: {smx_cols}')
            smx_cols=[x.lower() for x in smx_cols]
            print(smx_cols)
            df_tmp=smx_df[smx_cols]
            #print(df_tmp)
            print(Fore.CYAN+ 'after printing df_tmp')
            print(len(smx_cols))
            #print(row['prefix'])
            if len(smx_cols) > 1:
                df_tmp['concatenation']= df_tmp[df_tmp.columns].astype(str).agg('_'.join, axis=1)
                print(Fore.YELLOW+'concatenated values')
                print(df_tmp['concatenation'])
                print(Fore.CYAN+f'{process_type_str} and {len(process_type_str)}')
                
                df_tmp[row['parameter_name']]=df_tmp['concatenation'].values
                #if '21' in process_type_str:
                if key_type in ['BKEY_CALL', 'REG_BKEY_PROCESS']:
                    print(Fore.YELLOW+'value of process type is 21')
                    print(row['parameter_name'])
                    print(df_tmp['concatenation'].values)
                    la_liste=df_tmp['concatenation'].values
                    la_liste=['BK_'+x for x in la_liste]
                    print(la_liste)
   
            else:
                print(df_tmp)
                print(df_tmp.columns)
                print(df_tmp[smx_cols[0]])
                df_tmp[row['parameter_name']] = df_tmp[smx_cols[0]].values

            print(Fore.LIGHTMAGENTA_EX+ f'we need to apply the prefix if needed : {row['prefix']}')

            if pd.notna(row['prefix']): 
                print(Fore.LIGHTMAGENTA_EX+ f'we need to apply the prefix {row['prefix']}')
                la_liste=[f'{row['prefix']}_{x}' for x in df_tmp[row['parameter_name']]]
                df_tmp[row['parameter_name']]=la_liste
            if row['presentation_col']=='quoted':
                  print(Fore.MAGENTA+ f'we need to apply the quoted  {row['presentation_col']}')
                  print(row['parameter_name'])
                  df_tmp[row['parameter_name']]=df_tmp[row['parameter_name']].apply(lambda x: '\'{}\''.format(x))
                  #df_tmp[row['parameter_name']]=df_tmp[row['parameter_name']].apply(lambda x: f"'{x}'")
                  #df['col1'] = df['col1'].apply(lambda x: f"'{x}'")
                  #df['col1'] = "'" + df['col1'].astype(str) + "'"
                  #df_tmp[row['parameter_name']]="'"+df_tmp[row['parameter_name']].astype(str) + "'"



            print(row['parameter_name'])
            final_df[row['parameter_name']]=df_tmp[row['parameter_name']].values



    
    print(final_df)
    print(Fore.YELLOW+ f'printing final df type : {type(final_df)}')
    script = final_df.values.tolist()
    '''
    with open('REG_BKEY_PROCESS_output.SQL', mode='wt', encoding='utf-8') as myfile:
                for l in script:
                    myfile.write('\n'.join(l))
                    myfile.write('\n\n')
    '''
    print(Fore.YELLOW+" &&&&&&&&&&&&&&&&&   we end here ")
    print(Back.YELLOW+ "*********************************************************************************")
    '''
    end of new section
    '''

    
    #return df_with_empty_cols
    return final_df


def get_bkey_reg_script(smx_model, filtered_script_df, env_attributes, key_type):
    scripts=[]
    print(key_type)
    print(Fore.LIGHTRED_EX+" 0- Lets check the filtered operation *****************************************  :")
    print(filtered_script_df)

    print(filtered_script_df.columns)
    script=" place holder "
    print(env_attributes)
    print(type(env_attributes))
    
    print("==========================================================================")
   
    print(filtered_script_df)
    df_by_function=filtered_script_df.groupby(['operation', 'schema','functions' ], sort=False)
       
    for group_name, df_group in df_by_function:
        print(Back.RED + f"Group Name: {group_name}")
        print(Back.GREEN + str(group_name))
        e_schema = df_group['schema'].unique()
        print(e_schema)
        env_schema = env_attributes[e_schema[0]]
        print(env_schema)
        print(df_group.columns)
        print(df_group.shape)
        df_group['schema']=env_schema
        print(df_group['schema'])
        print(df_group.columns)
        print(Fore.LIGHTRED_EX+'1) ======================== apply schema and other environment variables ==========')
    
        print(df_group[['parameter_name','source', 'smx_column']])

        #df_group['parameter'] = df_group.apply(lambda row: str(env_attributes[row.parameter_name]).strip("\'") if  row.source=='env' else np.nan, axis=1)
        print(df_group.columns)
        
        smx_tab= df_group['source'].unique()
        print(df_group.columns)
        df_group['parameter'] = df_group.apply(lambda row: f'"{str(env_attributes[row.env_variable])}"' if  row.source=='env' else row.smx_column, axis=1)
        print(df_group)
        print(Back.CYAN + str(df_group['parameter']))

        
        print(Fore.LIGHTRED_EX+"2)  ==== Now we select the important columns, namely operation, schema and functions")
        #df=df_group[['operation','schema' ,'functions', 'source', 'smx_column', 'parameters', 'parameter','presentation_col']]
        df=df_group
        print(df.columns)
        print(df.shape)
        print(df)
        
        print(" oooooooooooooooooooooooooooooooooooooo join the parameters in one string tuple")
        print(" oooooooooooooooooooooooooooooooooooooo to be done after retrieving the values from the smx file ")

        parameters_string =tuple(list(df_group['parameter']))         
        print("printing parameters_string")
        print(parameters_string)

        print(type(parameters_string))
        print("assigning the same value to all rows")
        
        print(df)

        print(Fore.LIGHTRED_EX+"3) ====================== Now drop duplicates ")
        df=df.drop_duplicates()
        print(df)
        print(df.shape)


        print(Fore.LIGHTRED_EX+"4) ================================== aggregations operation with schema.function ")

        df['schema_functions'] = df[['schema', 'functions']].agg('.'.join, axis=1)
        print(df)
        df['operation_schema_functions'] = df[['operation','schema_functions']].agg(' '.join, axis=1)
        df= df.drop(['operation', 'schema', 'functions', 'schema_functions'], axis=1)
        print(df)


       
        print(Fore.LIGHTRED_EX+"5) preparing the list of smx tabs : removing env from source ")
        print(smx_tab)
        print(smx_tab[0])
        smx_tab_lst= list(smx_tab)
        if 'env' in list(smx_tab):
            smx_tab_lst.remove('env')
        
        
        
        smx_filtered_df = df[df['source'] != 'env']
        print(df)
        print(list(smx_filtered_df['source'].unique()))
        print(smx_filtered_df['smx_column'])
        print(list(smx_filtered_df['smx_column']))

       

        
        source_params_df= df[[ 'parameter_name', 'source', 'parameter', 'smx_column','env_variable','parameters','presentation_col','prefix', 'join_key']]
        print(Fore.BLUE+ " hhhh PRINTING source_params_df columns hhhhhhhh")
        print(source_params_df.columns)
        print(source_params_df)
        print(df.columns)
        print(Fore.LIGHTGREEN_EX+ f'smx_tab_list = {smx_tab_lst}')
        
        params_df = get_params_values_better(smx_tab_lst[0], source_params_df, smx_model, key_type)

        #params_df = get_smx_values(smx_tab[0],list(smx_filtered_df['smx_column']), smx_model)
        
        print(Back.LIGHTRED_EX + "====================== after get  values ============================")
        print(params_df)
       

        print(Back.CYAN + "params_df with combined parameters")
        cols=params_df.columns
        print(cols)

        merged_column = params_df.apply(lambda row: ', '.join(map(str, row)), axis=1)
        params_df['merged']= merged_column
        print(Back.CYAN +" ================== params_df[merged] --------------- ")
        print(Back.CYAN +  params_df['merged'])
        print( "params_df columns " + params_df.columns)
        print(Back.CYAN +" ================== df --------------- ")

        print(df)
        print(df.columns)
        print(df.shape)
        #df3= pd.merge(df, params_df, left_index=True, right_index=True, sort=True)
        #print(df3)
        #df3 = pd.concat([df, params_df], ignore_index=True)
        #print(df3)
        print(Back.CYAN +" ================== params_df --------------- ")

        #print(params_df)
        print(params_df.shape)
        print(params_df.columns)
        print(params_df['merged'])
        #df = pd.concat([df, params_df], ignore_index=True)
        df = df.reset_index()
        params_df=params_df.reset_index(drop=True)
        duplicate_cols = params_df.columns[params_df.columns.duplicated()]
        params_df.drop(columns=duplicate_cols, inplace=True)
        #df = pd.concat([df, params_df], ignore_index=True)
        print(Back.CYAN +" ================== check the number of cols  --------------- ")

        print(len(df.columns) == len(set(df.columns)))

        print(len(params_df.columns) == len(set(params_df.columns)))
        df = df.reset_index(drop=True)
        df = pd.concat([df, params_df], ignore_index=True)
        #print(df)
        cols=params_df.columns
        #params_df['combined'] = params_df[cols].apply(lambda row: ','.join(row.values.astype(str)), axis=1)
        #params_df['combined'] = params_df[cols].apply(lambda row: tuple(row.values.astype(str)), axis=1)
        print(Back.CYAN+ "df['operation_schema_functions'].unique()")
        function = df['operation_schema_functions'].unique()
        print(function)
        #params_df =params_df.assign(['function'], function[0])
        #params_df = params_df.assign(function=function)
        params_df['fn'] =function[0]
        params_df['combined']=params_df['merged'].apply(lambda row: f"({row})")
        params_df=params_df[['fn', 'combined']]
        print(Back.CYAN + "params_df['combined']")
        
        print(params_df['combined'])
        params_df['script']=params_df['fn']+params_df['combined']
        #print(Back.CYAN + "params_df['script']")

        #print(params_df['script'])
        fn_scripts=set(params_df['script'])
        print(Fore.LIGHTMAGENTA_EX+ "printing the fn_scripts")
        print(len(fn_scripts))
        fn_scripts = [item + ";" for item in fn_scripts]
        print(fn_scripts)
        scripts.append(list(fn_scripts))
        
        #break

    print(type(params_df['script']))

    for s in list(fn_scripts):
        print(Back.LIGHTRED_EX+ s)
   
    
    print(type(scripts))

    print(Back.CYAN+ str(scripts))
    
    print(type(scripts))
    
    print(len(scripts))
    print(scripts)

    return scripts

def smx_preprocess(smx_model, smx_tab, condition):
    print(Back.MAGENTA+" In smx preprocessing ----------")
    df = smx_model[smx_tab]
    print(df)
    print(type(smx_model))
    print(Back.MAGENTA+"condition")
    print(condition)
   
    print(df.columns)
   
    print(Fore.YELLOW+ condition)
    c = f"df[df{condition}]"
    print(c)
    result = eval(c)
    print(result)

    smx_model[smx_tab]=result
    return smx_model
  
def smx_pre_filter(smx_model, smx_tab, condition):
    print(smx_model)
    print(smx_tab)
    print(condition)
    df= smx_model[smx_tab]
    print(df)
    

    filtered_df= df[(df['key set name']=='PARTY') & (df['key domain name']=='CIF')]
    print(filtered_df)
    smx_model[smx_tab]=filtered_df
    df2=smx_model['stg tables']
    filtered_df= df2[df2['table name stg']=='STG_CB_CUSTOMER']
    smx_model['stg tables']=filtered_df
    
    return smx_model


def get_core_script_dict(script, smx_model):
        print(Fore.YELLOW+"=========================== ca commence =========================")

        df = smx_model['core tables']
        
        list_tables=df['table name'].unique()
        #print(list_tables)
        script_dict={}
        script_dict = dict.fromkeys(list_tables, [])
        #print(script_dict)

        #print(script)
        #print(Fore.YELLOW+f'{len(script)}')
        flat_list = [item for sublist in script for item in sublist]
        print(flat_list)
        print(Fore.YELLOW+f'{len(flat_list)}')

        for s in flat_list:
            #print(Fore.LIGHTGREEN_EX+ s)
            sub_s=  s[s.index("("):s.index(")")+1]
            print(Fore.YELLOW+sub_s)
            sub_s.replace('"', '')
            print(type(sub_s))
            print(Fore.MAGENTA+sub_s)
            my_tuple = eval(sub_s)
            #print(my_tuple)
            #print(type(my_tuple))
            #print(Fore.LIGHTRED_EX+'ss =  '+ my_tuple[1])
            ss=my_tuple[1]
            print(Fore.CYAN+f'{ss}')
            items=[]
            for k in list_tables:
                #print(Fore.BLUE+" "*4+'k = '+f'{k}')
                
                if k == ss:
                    print(Fore.BLUE+" "*4+f'{k}')
                    print(" "*4+ss)
                    print(" "*4+s)
                    print(" "*4+ f'{script_dict[k]}')
                    if script_dict[k] :
                        print(Fore.MAGENTA+f'{type(script_dict[k])}')
                        if isinstance(script_dict[k], list):
                            items2=script_dict[k]
                        else:
                            items2=[script_dict[k]]

                        print(Fore.YELLOW+ f'{items2}')
                        print(Fore.YELLOW+f'{type(items2)}')
                        
                        
                        items2.append(s)
                        items=items2
                        print(items)
                        script_dict.update({k:items})

                    else:#script_dict[k].append(s)
                        script_dict.update({k: s})
                        print(Fore.BLUE + f'In else ')
                        print(s)
                        print(type(s))

                    

                    print(" "*4+ f'{script_dict[k]}')

        print(Fore.YELLOW+"=========================== c fini =========================")
        for k in script_dict.keys():
            print(k)
            print(script_dict[k])
        print(len(script_dict))
        return script_dict



@st.cache_resource
def load_cached_model():
    script_file = "schema_functions_MAPPED_script_V_7_1.xlsx"
    return load_syntax_model(script_file)


def main(smx_model, key_type, env , bigint_flag):
    syntax_model = load_cached_model()
    
    filtered_script_df=filter_key_type(syntax_model, key_type)
    
    env_attributes=get_env_dict(str(env), bigint_flag)

    print('=====================================================================')
    #print(filtered_script_df)
    print(filtered_script_df.columns)
    print('=====================================================================')


    match key_type:
        case "BKEY_CALL" : 
             cols_list=['operation','schema','functions', 'parameters', 'parameter_name', 'source', 'smx_column'] 
             script= get_bkey_reg_script(smx_model, filtered_script_df, env_attributes, key_type)
            
        case "REG_BKEY_PROCESS":
            print("key type = ", key_type)
            print(filtered_script_df.columns)
            condition = "['Key Set Name' == 'PARTY' AND 'Key Domain Name' == 'CIF']"
            
            smx_model= smx_pre_filter(smx_model, 'bkey', condition)
            script= get_bkey_reg_script(smx_model, filtered_script_df, env_attributes, key_type)
                    
        case "REG_BKEY_DOMAIN" : #Stream".lower():
           print("key type = ", key_type)
           #script= get_bkey_domain_script(smx_model,filtered_script_df, env_attributes )
           script= get_bkey_reg_script(smx_model, filtered_script_df, env_attributes, key_type)

           print("======================= returned script : ")
           print(script)
           print(len(script))
           print(type(script))
           script= get_bkey_reg_script(smx_model, filtered_script_df, env_attributes, key_type)

        case "REG_BKEY" :# STG tables".lower():
            print("key type = ", key_type)
            print(env_attributes)
            script= get_bkey_reg_script(smx_model,filtered_script_df, env_attributes , key_type)
            #script= get_bkey_call_script(smx_model,filtered_script_df, env_attributes )
            print("======================= returned script : ")
            print(script)
            for s in script:
                print(Back.LIGHTGREEN_EX+ str(s))
            script= get_bkey_reg_script(smx_model,filtered_script_df, env_attributes , key_type)        
        case "STREAM" :# STG tables".lower():
            print("key type = ", key_type)
            print(env_attributes)
            script= get_bkey_reg_script(smx_model,filtered_script_df, env_attributes, key_type )
            #script= get_bkey_call_script(smx_model,filtered_script_df, env_attributes )
            print("======================= returned script : ")
            print(script)
            for s in script:
                print(Back.LIGHTGREEN_EX+ str(s))
            script= get_bkey_reg_script(smx_model,filtered_script_df, env_attributes, key_type )

        case "REG_BMAP" :# STG tables".lower():
            print("key type = ", key_type)
            print(env_attributes)
            script= get_bkey_reg_script(smx_model,filtered_script_df, env_attributes, key_type )
            #script= get_bkey_call_script(smx_model,filtered_script_df, env_attributes )
            print("======================= returned script : ")
            print(script)
            for s in script:
                print(Back.LIGHTGREEN_EX+ str(s))
            script= get_bkey_reg_script(smx_model,filtered_script_df, env_attributes, key_type )

        case "REG_BMAP_DOMAIN" :# STG tables".lower():
            print("key type = ", key_type)
            print(env_attributes)
            script= get_bkey_reg_script(smx_model,filtered_script_df, env_attributes, key_type )
            #script= get_bkey_call_script(smx_model,filtered_script_df, env_attributes )
            print("======================= returned script : ")
            print(script)
            for s in script:
                print(Back.LIGHTGREEN_EX+ str(s))
            
            script= get_bkey_reg_script(smx_model,filtered_script_df, env_attributes, key_type )
            
        case "EXEC_SRCI" :# STG tables".lower():
            print("key type = ", key_type)
            print(env_attributes)
            print(Fore.MAGENTA + " checking values in environment ")
            print(env_attributes.values)
            for index, value in enumerate(env_attributes.values()):
                print(f"Index: {index}, Value: {value}")
            script= get_bkey_reg_script(smx_model,filtered_script_df, env_attributes, key_type )          
            #script= get_bkey_call_script(smx_model,filtered_script_df, env_attributes )
            print("======================= returned script : ")
            print(script)
            print(Fore.CYAN+ " Checking the type of the script")
            print(type(script))
            for s in script:
                print(Back.LIGHTGREEN_EX+ str(s))

        case "CORE_KEY_COL_REG" :# STG tables".lower():
            print("key type = ", key_type)
            print(env_attributes)
            condition = "['pk'] == 'Y'"
            smx_model= smx_preprocess(smx_model, 'core tables', condition)
            script= get_bkey_reg_script(smx_model,filtered_script_df, env_attributes, key_type )
            script_dict= get_core_script_dict(script, smx_model)

            script2 = ''' '''  # Initialize with triple quotes for multiline string
            for table_name, values in script_dict.items():
                print(table_name)
                print(values)
                stmnt = f"""
SELECT * FROM G{env}1V_GCFR.GCFR_TRANSFORM_KEYCOL WHERE OUT_OBJECT_NAME = '{table_name}'AND OUT_DB_NAME = 'G{env}1V_CORE';
DELETE FROM G{env}1V_GCFR.GCFR_TRANSFORM_KEYCOL WHERE OUT_OBJECT_NAME = '{table_name}'AND OUT_DB_NAME = 'G{env}1V_CORE';
                    """
                if isinstance(values, list):
                    script2 += stmnt
                    # Then append all values
                    for value in values:
                        script2 += value + '\n'
                else:
                    script2 += stmnt
                    # Then append the single value
                    script2 += values + '\n'
            return script2  # Return the combined script
        case "REG_CORE_PROCESS" :# core tables
            print("key type = ", key_type)
            print(env_attributes)
            #smx_model= smx_preprocess(smx_model, 'core tables', f'PK==Y')
            script= get_bkey_reg_script(smx_model,filtered_script_df, env_attributes, key_type )

        case "HIST_REG" :# STG tables".lower():
            print("key type = ", key_type)
            print(env_attributes)
            
            condition= "['historization key'].notnull()"
            smx_model= smx_preprocess(smx_model, 'core tables', condition)
            print(filtered_script_df.columns)
            print(filtered_script_df['preprocessing'].dropna().unique())
            conditions= filtered_script_df['preprocessing'].dropna().unique()
            print(conditions)
            #sys.exit(0)
            script= get_bkey_reg_script(smx_model,filtered_script_df, env_attributes, key_type )
            
            print("======================= returned script : ")
            print(script)
            for s in script:
                print(Back.LIGHTGREEN_EX+ str(s))
            
            print(type(smx_model))
            print(smx_model.keys())
            print(smx_model['core tables'])
            
            script_dict= get_core_script_dict(script, smx_model)
            stmnt = """SELECT * FROM G{env}1V_GCFR.GCFR_TRANSFORM_HISTCOL WHERE OUT_OBJECT_NAME = '{table_name}' AND OUT_DB_NAME = 'G{env}1V_CORE';
DELETE FROM G{env}1V_GCFR.GCFR_TRANSFORM_HISTCOL WHERE OUT_OBJECT_NAME = '{table_name}' AND OUT_DB_NAME = 'G{env}1V_CORE';  """
            script = df_utlis.add_sql_to_dictionary(script_dict,env,stmnt)
        case "bkey_views":
            script = Queries.generate_bkey_views(smx_model,env)
        case "Insert BMAP values":
            script = Queries.insert_bmap_values(smx_model,env)
        case"Create LKP views":
            script = Queries.create_LKP_views (smx_model, env)
        case"create_stg_table_and_view":
            script = Queries.create_stg_table_and_view (smx_model, env)
        case"create_SCRI_table":
            script = Queries.create_SCRI_table (smx_model, env)
        case"create_SCRI_view":
            script = Queries.create_SCRI_view (smx_model, env)
        case"create_SCRI_input_view":
           script = Queries.create_SCRI_input_view (smx_model, env)
        case"create_core_table":
            script = Queries.create_core_table (smx_model, env)
        case"create_core_table_view":
            script = Queries.create_core_table_view (smx_model, env)
        case"create_core_input_view":
            script = Queries.create_core_input_view (smx_model, env)
            
    return script

    

    