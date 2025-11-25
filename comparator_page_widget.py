# comparison_page_widget.py
import sys
import pandas as pd
import openpyxl
import difflib
from PySide6.QtCore import (
    QAbstractTableModel,
    QModelIndex,
    Qt,
    Slot,
)
from PySide6.QtWidgets import (
    QApplication,
    QWidget,  # Changed from QMainWindow
    QVBoxLayout,
    QHBoxLayout,
    QPushButton,
    QLabel,
    QFileDialog,
    QTableView,
    QComboBox,
    QMessageBox,
    QHeaderView,
)

# --- Authentication (Stub) ---
# This class will be created by main_app.py and passed in
class Authenticator:
    def logout(self):
        print("Logout requested (stub)")
        QApplication.instance().quit()


# --- Pandas DataFrame Model for QTableView ---
class PandasModel(QAbstractTableModel):
    def __init__(self, data=pd.DataFrame()):
        super().__init__()
        self._data = data

    def rowCount(self, parent=QModelIndex()):
        return self._data.shape[0]

    def columnCount(self, parent=QModelIndex()):
        return self._data.shape[1]

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None
        if role == Qt.ItemDataRole.DisplayRole:
            return str(self._data.iloc[index.row(), index.column()])
        return None

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.DisplayRole:
            if orientation == Qt.Orientation.Horizontal:
                return str(self._data.columns[section])
            if orientation == Qt.Orientation.Vertical:
                return str(self._data.index[section])
        return None
    
    def setData(self, data):
        self.beginResetModel()
        self._data = data
        self.endResetModel()


# --- Main Comparator Page Widget ---
class ComparatorPageWidget(QWidget): # <-- RENAMED AND CHANGED
    def __init__(self, authenticator, parent=None):
        super().__init__(parent)
        self.authenticator = authenticator

        # --- Store workbook data ---
        self.wb1_sheets = None
        self.wb2_sheets = None
        self.wb1_path = ""
        self.wb2_path = ""
        
        # --- Main Layout (is now a horizontal layout) ---
        main_layout = QHBoxLayout(self)

        # --- Create Sidebar ---
        sidebar_widget = self._create_sidebar()

        # --- Create Main Panel (this was the old central_widget) ---
        main_panel = QWidget()
        main_panel_layout = QVBoxLayout(main_panel)

        # 1. File Upload Area
        self._create_upload_ui(main_panel_layout)
        # 2. Results Area
        self._create_results_ui(main_panel_layout)
        
        # 3. Status/Info Label
        self.info_label = QLabel("Upload two WAVZ SMX workbooks to compare them")
        self.info_label.setStyleSheet("font-style: italic; color: gray;")
        main_panel_layout.addWidget(self.info_label)

        # Add sidebar and main panel to the page's layout
        main_layout.addWidget(sidebar_widget, 1) # 1-part stretch
        main_layout.addWidget(main_panel, 4)     # 4-parts stretch

    def _create_sidebar(self):
        # --- REWRITTEN from QDockWidget to QWidget ---
        widget = QWidget()
        layout = QVBoxLayout(widget)
        widget.setObjectName("sidebar")
        widget.setStyleSheet("#sidebar { background-color: #f0f2f6; }")

        self.logout_button = QPushButton("Logout")
        # The 'clicked' signal will be connected in main_app.py
        
        layout.addWidget(self.logout_button)
        layout.addStretch()  # Pushes button to the top
        
        return widget

    def _create_upload_ui(self, main_layout):
        upload_layout = QHBoxLayout()

        # --- Column 1: Workbook 1 ---
        col1_layout = QVBoxLayout()
        col1_label = QLabel("Workbook 1")
        col1_label.setStyleSheet("font-size: 16px; font-weight: bold;")
        self.wb1_button = QPushButton("Upload Workbook 1")
        self.wb1_label = QLabel("No file selected.")
        
        col1_layout.addWidget(col1_label)
        col1_layout.addWidget(self.wb1_button)
        col1_layout.addWidget(self.wb1_label)
        
        # --- Column 2: Workbook 2 ---
        col2_layout = QVBoxLayout()
        col2_label = QLabel("Workbook 2")
        col2_label.setStyleSheet("font-size: 16px; font-weight: bold;")
        self.wb2_button = QPushButton("Upload Workbook 2")
        self.wb2_label = QLabel("No file selected.")

        col2_layout.addWidget(col2_label)
        col2_layout.addWidget(self.wb2_button)
        col2_layout.addWidget(self.wb2_label)
        
        upload_layout.addLayout(col1_layout)
        upload_layout.addLayout(col2_layout)
        main_layout.addLayout(upload_layout)
        
        # --- Connect signals to slots ---
        self.wb1_button.clicked.connect(self.on_upload_wb1)
        self.wb2_button.clicked.connect(self.on_upload_wb2)

    def _create_results_ui(self, main_layout):
        # --- Models ---
        self.unique_sheets_model = PandasModel()
        self.only_in_df1_model = PandasModel()
        self.only_in_df2_model = PandasModel()
        self.different_rows_model = PandasModel()

        # --- Unique Sheets ---
        main_layout.addWidget(QLabel("📄 Sheets Present in Only One Workbook"))
        self.unique_sheets_view = QTableView()
        self.unique_sheets_view.setModel(self.unique_sheets_model)
        main_layout.addWidget(self.unique_sheets_view)

        # --- Drill Down ---
        main_layout.addWidget(QLabel("🔍 Drill Down Analysis"))
        self.drill_down_combo = QComboBox()
        main_layout.addWidget(self.drill_down_combo)
        
        # --- Drill Down Results (in a horizontal layout) ---
        drill_down_layout = QHBoxLayout()
        
        # Only in WB1
        v1_layout = QVBoxLayout()
        v1_layout.addWidget(QLabel("Rows only in Workbook 1"))
        self.only_in_df1_view = QTableView()
        self.only_in_df1_view.setModel(self.only_in_df1_model)
        v1_layout.addWidget(self.only_in_df1_view)
        
        # Only in WB2
        v2_layout = QVBoxLayout()
        v2_layout.addWidget(QLabel("Rows only in Workbook 2"))
        self.only_in_df2_view = QTableView()
        self.only_in_df2_view.setModel(self.only_in_df2_model)
        v2_layout.addWidget(self.only_in_df2_view)

        # Different Rows
        v3_layout = QVBoxLayout()
        v3_layout.addWidget(QLabel("Rows with differences"))
        self.different_rows_view = QTableView()
        self.different_rows_view.setModel(self.different_rows_model)
        v3_layout.addWidget(self.different_rows_view)
        
        drill_down_layout.addLayout(v1_layout)
        drill_down_layout.addLayout(v2_layout)
        drill_down_layout.addLayout(v3_layout)
        main_layout.addLayout(drill_down_layout)
        
        # --- Connect signal ---
        self.drill_down_combo.currentIndexChanged.connect(self.on_drill_down)
        
    def _resize_table(self, table_view):
        table_view.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)

    # --- SLOTS (These are the event handlers) ---
    @Slot()
    def on_upload_wb1(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Upload Workbook 1", "", "Excel Files (*.xlsx)")
        if file_path:
            self.wb1_path = file_path
            self.wb1_label.setText(file_path.split('/')[-1])
            self.info_label.setText("Loading Workbook 1...")
            QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
            self.wb1_sheets = self.load_workbook(self.wb1_path)
            QApplication.restoreOverrideCursor()
            self.info_label.setText("Loaded Workbook 1.")
            self.run_comparison()

    @Slot()
    def on_upload_wb2(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Upload Workbook 2", "", "Excel Files (*.xlsx)")
        if file_path:
            self.wb2_path = file_path
            self.wb2_label.setText(file_path.split('/')[-1])
            self.info_label.setText("Loading Workbook 2...")
            QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
            self.wb2_sheets = self.load_workbook(self.wb2_path)
            QApplication.restoreOverrideCursor()
            self.info_label.setText("Loaded Workbook 2.")
            self.run_comparison()
            
    def run_comparison(self):
        if not self.wb1_sheets or not self.wb2_sheets:
            self.info_label.setText("Please upload both workbooks to compare.")
            return
        
        self.info_label.setText("Analyzing all sheets...")
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        
        summary_df = self.compare_full_workbooks(self.wb1_sheets, self.wb2_sheets)
        
        unique_sheets = summary_df[summary_df['Status'].isin(['Only in Workbook 1', 'Only in Workbook 2'])]
        self.unique_sheets_model.setData(unique_sheets)
        self._resize_table(self.unique_sheets_view)

        self.drill_down_combo.clear()
        sheets_with_differences = summary_df[summary_df['Status'] == 'Has Differences']
        if not sheets_with_differences.empty:
            self.drill_down_combo.addItem("Select a sheet...")
            self.drill_down_combo.addItems(sheets_with_differences['Sheet'].tolist())
        
        QApplication.restoreOverrideCursor()
        self.info_label.setText("Comparison complete.")

    @Slot()
    def on_drill_down(self, index):
        sheet_name = self.drill_down_combo.currentText()
        if not sheet_name or sheet_name == "Select a sheet...":
            self.only_in_df1_model.setData(pd.DataFrame())
            self.only_in_df2_model.setData(pd.DataFrame())
            self.different_rows_model.setData(pd.DataFrame())
            return
            
        if self.wb1_sheets and self.wb2_sheets:
            df1 = self.wb1_sheets[sheet_name]
            df2 = self.wb2_sheets[sheet_name]
            
            only_in_df1, only_in_df2, different_rows = self.compare_rows(df1, df2)
            
            self.only_in_df1_model.setData(only_in_df1)
            self.only_in_df2_model.setData(only_in_df2)
            self.different_rows_model.setData(different_rows)

            self._resize_table(self.only_in_df1_view)
            self._resize_table(self.only_in_df2_view)
            self._resize_table(self.different_rows_view)

    # --- Backend Logic (Copied directly from your script) ---
    def load_workbook(self, uploaded_file):
        if uploaded_file is None:
            return None
        xls = pd.ExcelFile(uploaded_file)
        sheets = {}
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name, header=0)
            df = df.dropna(how='all').dropna(axis=1, how='all')
            sheets[sheet_name] = df.copy() # <- THE FIX
        return sheets

    def compare_rows(self, df1, df2):
        if df1 is None or df2 is None:
            return None, None, None
        OuterJoinedDF = pd.merge(df1, df2, how="outer", indicator="Exist")
        only_in_df1 = OuterJoinedDF.query("Exist == 'left_only'")
        only_in_df2 = OuterJoinedDF.query("Exist == 'right_only'")
        different_rows_df = OuterJoinedDF.query("Exist != 'both'")
        return only_in_df1, only_in_df2, different_rows_df

    def compare_full_workbooks(self, wb1_sheets, wb2_sheets):
        comparison_summary = []
        all_sheets = set(wb1_sheets.keys()) | set(wb2_sheets.keys())
        for sheet_name in sorted(all_sheets):
            sheet_summary = {
                'Sheet': sheet_name, 'Status': '', 'Rows Only in WB1': 0,
                'Rows Only in WB2': 0, 'Rows with Differences': 0
            }
            if sheet_name not in wb1_sheets:
                sheet_summary['Status'] = 'Only in Workbook 2'
            elif sheet_name not in wb2_sheets:
                sheet_summary['Status'] = 'Only in Workbook 1'
            else:
                df1 = wb1_sheets[sheet_name]
                df2 = wb2_sheets[sheet_name]
                if len(df1.columns) > 0 and len(df2.columns) > 0:
                    only_in_df1, only_in_df2, different_rows = self.compare_rows(df1, df2)
                    sheet_summary['Rows Only in WB1'] = len(only_in_df1)
                    sheet_summary['Rows Only in WB2'] = len(only_in_df2)
                    sheet_summary['Rows with Differences'] = len(different_rows) if different_rows is not None else 0
                    total_diffs = sheet_summary['Rows Only in WB1'] + sheet_summary['Rows Only in WB2'] + sheet_summary['Rows with Differences']
                    sheet_summary['Status'] = 'Identical' if total_diffs == 0 else 'Has Differences'
                else:
                    sheet_summary['Status'] = 'Cannot Compare'
            comparison_summary.append(sheet_summary)
        return pd.DataFrame(comparison_summary)

